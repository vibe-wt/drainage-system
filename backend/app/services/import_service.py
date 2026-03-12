from __future__ import annotations

from copy import deepcopy
from datetime import datetime
from io import BytesIO
import json
from uuid import uuid4

from fastapi import HTTPException, UploadFile
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.schemas.imports import CreateImportBatchRequest
from app.schemas.manhole import CreateManholeRequest
from app.schemas.pipe import CreatePipeRequest
from app.services.manhole_service import create_manhole
from app.services.pipe_service import create_pipe

IMPORT_BATCHES: dict[str, dict] = {}


def _now_iso() -> str:
    return datetime.now().isoformat(timespec="seconds")


def _serialize_batch(batch: dict) -> dict:
    return {
        "id": batch["id"],
        "batch_name": batch["batch_name"],
        "source_type": batch["source_type"],
        "object_type": batch["object_type"],
        "import_status": batch["import_status"],
        "file_name": batch.get("file_name"),
        "total_count": batch.get("total_count", 0),
        "success_count": batch.get("success_count", 0),
        "failed_count": batch.get("failed_count", 0),
        "created_at": batch["created_at"],
        "preview_rows": deepcopy(batch.get("preview_rows", [])),
        "columns": list(batch.get("columns", [])),
        "error_summary": list(batch.get("error_summary", [])),
        "imported_objects": deepcopy(batch.get("imported_objects", [])),
    }


def list_import_batches() -> list[dict]:
    batches = sorted(IMPORT_BATCHES.values(), key=lambda item: item["created_at"], reverse=True)
    return [_serialize_batch(batch) for batch in batches]


def create_import_batch(payload: CreateImportBatchRequest) -> dict:
    batch_id = f"batch_{uuid4().hex[:10]}"
    IMPORT_BATCHES[batch_id] = {
        "id": batch_id,
        "batch_name": payload.batch_name,
        "source_type": payload.source_type,
        "object_type": payload.object_type,
        "import_status": "created",
        "created_at": _now_iso(),
        "preview_rows": [],
        "columns": [],
        "records": [],
        "error_summary": [],
        "imported_objects": [],
    }
    return _serialize_batch(IMPORT_BATCHES[batch_id])


def get_import_batch(batch_id: str) -> dict:
    batch = IMPORT_BATCHES.get(batch_id)
    if batch is None:
        raise HTTPException(status_code=404, detail="导入批次不存在。")
    return batch


async def upload_import_file(batch_id: str, file: UploadFile) -> dict:
    batch = get_import_batch(batch_id)
    payload = await file.read()
    if not payload:
        raise HTTPException(status_code=400, detail="上传文件为空。")

    if batch["source_type"] == "excel":
        columns, records = _parse_excel(payload)
    elif batch["source_type"] == "geojson":
        columns, records = _parse_geojson(payload, batch["object_type"])
    else:
        raise HTTPException(status_code=400, detail="暂不支持的导入类型。")

    batch["file_name"] = file.filename
    batch["columns"] = columns
    batch["records"] = records
    batch["preview_rows"] = records[:5]
    batch["total_count"] = len(records)
    batch["import_status"] = "uploaded"
    batch["error_summary"] = []
    batch["imported_objects"] = []
    return _serialize_batch(batch)


def preview_import(batch_id: str, mapping: dict[str, str]) -> dict:
    batch = get_import_batch(batch_id)
    preview_rows = [_normalize_record(batch["object_type"], record, mapping) for record in batch["records"][:5]]
    return {
        "columns": list(batch["columns"]),
        "preview_rows": preview_rows,
        "detected_geometry_type": "Point" if batch["object_type"] == "manholes" else "LineString",
        "errors": [],
    }


def commit_import(db: Session, batch_id: str, mapping: dict[str, str]) -> dict:
    batch = get_import_batch(batch_id)
    success_count = 0
    failed_count = 0
    errors: list[str] = []
    imported_objects: list[dict[str, str]] = []

    for index, record in enumerate(batch["records"], start=1):
        try:
            normalized = _normalize_record(batch["object_type"], record, mapping)
            if batch["object_type"] == "manholes":
                payload = CreateManholeRequest(**normalized)
                created = create_manhole(db, payload)
            else:
                payload = CreatePipeRequest(**normalized)
                created = create_pipe(db, payload)
            success_count += 1
            imported_objects.append(
                {
                    "id": created["id"],
                    "code": created["code"],
                    "object_type": created["object_type"],
                }
            )
        except Exception as exc:
            failed_count += 1
            errors.append(f"第 {index} 行导入失败：{exc.detail if hasattr(exc, 'detail') else str(exc)}")

    batch["success_count"] = success_count
    batch["failed_count"] = failed_count
    batch["import_status"] = "completed" if failed_count == 0 else "completed_with_errors"
    batch["error_summary"] = errors[:10]
    batch["imported_objects"] = imported_objects
    return {
        "batch_id": batch_id,
        "total_count": len(batch["records"]),
        "success_count": success_count,
        "failed_count": failed_count,
        "errors": errors[:10],
        "imported_objects": imported_objects,
    }


def _parse_pipe_coordinates(value: object) -> list[list[float]]:
    if value is None:
        return []
    coordinates = value
    if isinstance(coordinates, str):
        coordinates = json.loads(coordinates)
    if not isinstance(coordinates, list):
        raise ValueError("coordinates 字段格式不正确")

    parsed: list[list[float]] = []
    for point in coordinates:
        if not isinstance(point, (list, tuple)) or len(point) < 2:
            raise ValueError("coordinates 需要是坐标数组")
        parsed.append([float(point[0]), float(point[1])])
    return parsed


def _parse_excel(payload: bytes) -> tuple[list[str], list[dict]]:
    workbook = load_workbook(filename=BytesIO(payload), read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="Excel 文件没有可读取内容。")
    headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    records: list[dict] = []
    for row in rows[1:]:
        if row is None or all(cell is None or str(cell).strip() == "" for cell in row):
            continue
        record = {headers[index]: row[index] for index in range(min(len(headers), len(row))) if headers[index]}
        records.append(record)
    return headers, records


def _parse_geojson(payload: bytes, object_type: str) -> tuple[list[str], list[dict]]:
    data = json.loads(payload.decode("utf-8"))
    features = data.get("features", [])
    records: list[dict] = []
    columns: set[str] = set()
    for feature in features:
        properties = feature.get("properties", {}) or {}
        geometry = feature.get("geometry", {}) or {}
        record = dict(properties)
        if object_type == "manholes":
            record["coordinates"] = geometry.get("coordinates", [])
        else:
            record["coordinates"] = geometry.get("coordinates", [])
        records.append(record)
        columns.update(record.keys())
    return sorted(columns), records


def _pick(record: dict, mapping: dict[str, str], field: str, fallback: str | None = None):
    source_key = mapping.get(field, fallback or field)
    return record.get(source_key) if source_key else None


def _normalize_record(object_type: str, record: dict, mapping: dict[str, str]) -> dict:
    if object_type == "manholes":
        coordinates = record.get("coordinates")
        if not coordinates:
            lng = _pick(record, mapping, "lng")
            lat = _pick(record, mapping, "lat")
            coordinates = [float(lng), float(lat)] if lng is not None and lat is not None else None
        return {
            "code": str(_pick(record, mapping, "code", "code") or ""),
            "name": str(_pick(record, mapping, "name", "name") or "未命名检查井"),
            "risk_level": str(_pick(record, mapping, "risk_level", "risk_level") or "low"),
            "manhole_type": str(_pick(record, mapping, "manhole_type", "manhole_type") or "污水井"),
            "catchment_name": str(_pick(record, mapping, "catchment_name", "catchment_name") or ""),
            "depth_m": float(_pick(record, mapping, "depth_m", "depth_m") or 0),
            "coordinates": [float(coordinates[0]), float(coordinates[1])],
        }

    coordinates = _parse_pipe_coordinates(record.get(mapping.get("coordinates", "coordinates")))
    return {
        "code": str(_pick(record, mapping, "code", "code") or ""),
        "name": str(_pick(record, mapping, "name", "name") or "未命名管道"),
        "risk_level": str(_pick(record, mapping, "risk_level", "risk_level") or "low"),
        "pipe_type": str(_pick(record, mapping, "pipe_type", "pipe_type") or "污水"),
        "diameter_mm": int(float(_pick(record, mapping, "diameter_mm", "diameter_mm") or 400)),
        "start_manhole_id": str(_pick(record, mapping, "start_manhole_id", "start_manhole_id") or ""),
        "end_manhole_id": str(_pick(record, mapping, "end_manhole_id", "end_manhole_id") or ""),
        "coordinates": coordinates,
    }
